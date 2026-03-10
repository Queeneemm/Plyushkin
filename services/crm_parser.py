import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from bot.utils.text import normalize_name


@dataclass(slots=True)
class CRMParserConfig:
    name_column: str
    stock_column: str
    header_row: int = 1
    sheet_name: str | None = None


class CRMExcelParser:
    def __init__(self, config: CRMParserConfig):
        self.config = config

    @staticmethod
    def _normalize_header(value: str) -> str:
        normalized = normalize_name(value).lower().replace('ё', 'е')
        return re.sub(r'[^a-zа-я0-9]+', '', normalized)

    def _get_headers_map(self, row) -> dict[str, int]:
        mapping = {}
        for idx, cell in enumerate(row, start=1):
            if cell.value:
                mapping[self._normalize_header(str(cell.value))] = idx
        return mapping

    def _find_column_indices(self, ws) -> tuple[int | None, int | None, int]:
        name_aliases = ('название', 'наименование')
        stock_aliases = ('остаток', 'остнаскладе', 'остатокнаскладе')

        def _resolve_indices(header_map: dict[str, int]) -> tuple[int | None, int | None]:
            name_idx = header_map.get(self._normalize_header(self.config.name_column))
            stock_idx = header_map.get(self._normalize_header(self.config.stock_column))

            if not name_idx:
                for alias in name_aliases:
                    name_idx = header_map.get(self._normalize_header(alias))
                    if name_idx:
                        break

            if not stock_idx:
                for alias in stock_aliases:
                    stock_idx = header_map.get(self._normalize_header(alias))
                    if stock_idx:
                        break

            return name_idx, stock_idx

        configured_row = self.config.header_row
        header_cells = next(ws.iter_rows(min_row=configured_row, max_row=configured_row))
        header_map = self._get_headers_map(header_cells)
        name_idx, stock_idx = _resolve_indices(header_map)
        if name_idx and stock_idx:
            return name_idx, stock_idx, configured_row

        scan_to_row = min(ws.max_row, configured_row + 10)
        for row_idx in range(1, scan_to_row + 1):
            if row_idx == configured_row:
                continue
            candidate_cells = next(ws.iter_rows(min_row=row_idx, max_row=row_idx))
            candidate_map = self._get_headers_map(candidate_cells)
            cand_name_idx, cand_stock_idx = _resolve_indices(candidate_map)
            if cand_name_idx and cand_stock_idx:
                return cand_name_idx, cand_stock_idx, row_idx

        return name_idx, stock_idx, configured_row

    def parse_stock(self, file_path: str | Path) -> dict[str, float]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[self.config.sheet_name] if self.config.sheet_name and self.config.sheet_name in wb.sheetnames else wb.active

        name_idx, stock_idx, header_row = self._find_column_indices(ws)

        if not name_idx or not stock_idx:
            raise ValueError('Не найдены требуемые колонки в CRM файле')

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=header_row + 1):
            raw_name = row[name_idx - 1].value
            raw_stock = row[stock_idx - 1].value
            if raw_name is None:
                continue
            name = normalize_name(str(raw_name))
            if not name:
                continue
            try:
                stock = float(raw_stock or 0)
            except (TypeError, ValueError):
                stock = 0
            result[name] = stock

        return result
